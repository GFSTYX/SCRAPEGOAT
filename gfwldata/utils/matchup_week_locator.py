import logging
import re

from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class MatchupWeekLocator:
    """Locates and calculates dimensions for weekly matchup data in a worksheet."""

    def __init__(self, matchups_sheet: Worksheet):
        self.matchups_sheet = matchups_sheet

    def locate_weeks_dimensions(
        self, wars_horizontal: int, wars_vertical: int, max_rows: int = 400
    ) -> list[dict]:
        """
        Locate week cells and calculate their corresponding war grid dimensions.

        Args:
            wars_horizontal: Number of wars across
            wars_vertical: Number of wars down
            max_rows: Maximum rows to search
        """
        logger.debug("Finding week cells in matchups sheet")

        week_cells = self._find_week_cells(max_rows)
        logger.debug("Found %i week cells", len(week_cells))

        week_data = []

        for week_cell in week_cells:
            war_dimensions = self._calculate_war_dimensions(
                week_cell, wars_horizontal, wars_vertical
            )
            week_data.append({"week_cell": week_cell, "war_dimensions": war_dimensions})

        return week_data

    def _find_week_cells(self, max_rows: int = 400) -> list[str]:
        """
        Find all cells in column A that exactly match 'Week N' pattern.

        Args:
            max_rows: Maximum number of rows to search
        """
        week_cells = []

        for row in range(1, max_rows + 1):
            cell_value = self.matchups_sheet.cell(row=row, column=1).value

            if cell_value and isinstance(cell_value, str):
                if re.match(r"^Week\s*(\d+)$", cell_value):
                    week_cells.append(f"A{row}")

        return sorted(week_cells)

    def _calculate_war_dimensions(
        self, week_cell: str, wars_horizontal: int = 4, wars_vertical: int = 4
    ) -> list[tuple[str, str]]:
        """
        Calculate war dimensions for a grid of wars based on a week cell.

        Args:
            week_cell: Cell reference for the week
            wars_horizontal: Number of wars horizontally
            wars_vertical: Number of wars vertically

        Returns:
            List of tuples containing (top_left, bottom_right) cell references for each war
        """
        # Convert week cell to coordinates
        week_coord = coordinate_to_tuple(week_cell)
        start_row, start_col = week_coord

        # War layout constants
        WAR_WIDTH = 3  # columns per war (A->C = 3 columns)
        WAR_HEIGHT = 6  # rows per war
        COL_OFFSET = 4  # each war starts 4 columns after the previous (A->E->I->M)
        ROW_SPACING = 1  # rows between wars
        WAR_START_ROW_OFFSET = 2  # First war starts 2 rows below week cell
        WAR_START_ROW = start_row + WAR_START_ROW_OFFSET

        dimensions = []

        # Calculate dimensions for each war in the grid
        for v in range(wars_vertical):
            for h in range(wars_horizontal):
                # Calculate top-left cell
                current_row = WAR_START_ROW + (v * (WAR_HEIGHT + ROW_SPACING))
                current_col = start_col + (h * COL_OFFSET)

                # Calculate bottom-right cell
                end_row = current_row + WAR_HEIGHT - 1
                end_col = current_col + WAR_WIDTH - 1

                # Convert to cell references
                top_left = f"{get_column_letter(current_col)}{current_row}"
                bottom_right = f"{get_column_letter(end_col)}{end_row}"

                dimensions.append((top_left, bottom_right))

        return dimensions
