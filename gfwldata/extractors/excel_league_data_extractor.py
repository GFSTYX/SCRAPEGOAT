import logging
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.workbook.workbook import Workbook

from gfwldata.config.worksheet import WorksheetSettings
from gfwldata.utils.matchup_week_locator import MatchupWeekLocator

logger = logging.getLogger(__name__)


class ExcelLeagueDataExtractor:
    """Extracts league data from an Excel file."""

    def __init__(self, config: WorksheetSettings):
        """Initializes the ExcelLeagueDataExtractor."""
        self.config = config
        self.workbook = self._load_workbook()
        self._initialize_sheets()

    def _load_workbook(self) -> Workbook:
        """Loads the workbook from the specified file path."""
        try:
            logger.info("Loading workbook from %s", self.config.FILEPATH)

            return load_workbook(self.config.FILEPATH, data_only=True)

        except Exception:
            logger.exception("Failed to load workbook from %s", self.config.FILEPATH)
            raise

    def _initialize_sheets(self) -> None:
        """Initializes the deck history and matchups sheets."""
        try:
            logger.info(
                "Loading sheets: %s, %s",
                self.config.DECK_HISTORY_SHEET_NAME,
                self.config.MATCHUPS_SHEET_NAME,
            )

            self.deck_history_sheet = self.workbook[self.config.DECK_HISTORY_SHEET_NAME]
            self.matchups_sheet = self.workbook[self.config.MATCHUPS_SHEET_NAME]

        except Exception:
            logger.exception("Failed to find worksheet")
            raise

    def get_league_data(self) -> tuple:
        """Extracts and returns matchups and deck history data."""
        logger.info("Starting league data extraction")

        matchups_data = self._get_matchups_data()
        deck_history_data = self._get_deck_history_data()

        logger.info("League data extraction complete")

        return matchups_data, deck_history_data

    def _get_deck_history_data(self) -> pd.DataFrame:
        """Extracts deck history data from the deck history sheet."""
        logger.info("Extracting deck history data")

        week_headers = self._get_week_headers()

        deck_history_data = []
        ROW_OFFSET = 2

        for row in range(ROW_OFFSET, self.config.DECK_HISTORY_MAX_ROWS):
            player_name = self.deck_history_sheet.cell(row=row, column=1).value
            if not player_name:
                continue

            for week_idx, week in enumerate(week_headers):
                deck_type = self.deck_history_sheet.cell(
                    row=row, column=week_idx + ROW_OFFSET
                ).value

                if deck_type:
                    week_number = int(week.split()[1])
                    deck_history_data.append(
                        {
                            "season": self.config.SEASON,
                            "week": week_number,
                            "player_name": player_name,
                            "deck_type": deck_type,
                        }
                    )

        return pd.DataFrame(deck_history_data)

    def _get_week_headers(self) -> list[str]:
        """Extracts week headers from the deck history sheet."""
        week_headers = []
        current_column = 2

        while True:
            cell_value = self.deck_history_sheet.cell(
                row=1, column=current_column
            ).value
            if not cell_value:
                break

            if isinstance(cell_value, str) and cell_value.startswith("Week"):
                week_headers.append(cell_value)

            current_column += 1

        return week_headers

    def _get_matchups_data(self) -> pd.DataFrame:
        """Extracts matchups data from the matchups sheet."""
        logger.info("Extracting matchups data")

        # Locate the dimensions of the wars for each week
        locator = MatchupWeekLocator(self.matchups_sheet)
        weeks_dimensions = locator.locate_weeks_dimensions(
            self.config.WARS_HORIZONTAL,
            self.config.WARS_VERTICAL,
            self.config.WEEK_SCAN_MAX_ROWS,
        )

        all_weeks_of_wars = []

        for week in weeks_dimensions:
            war_dimensions = week["war_dimensions"]
            all_weeks_of_wars.append(self._parse_week_of_wars_data(war_dimensions))

        if not all_weeks_of_wars:
            logger.error("No weeks of wars found in the matchups sheet.")
            return pd.DataFrame()

        # Combine all weeks of wars into a single DataFrame
        matchups_data = pd.concat(all_weeks_of_wars, ignore_index=True)

        return matchups_data

    def _parse_week_of_wars_data(self, war_dimensions: list[tuple]) -> pd.DataFrame:
        """Parses data for a week of wars."""
        week_of_wars = []

        for war_dimensions in war_dimensions:
            start_cell = war_dimensions[0]
            week_of_wars.append(self._parse_war_data(start_cell))

        return pd.concat(week_of_wars, ignore_index=True)

    def _parse_war_data(self, start_cell: str) -> pd.DataFrame:
        """Parses a single war's data."""
        # Convert the start cell to row and column indices
        start_coord = coordinate_to_tuple(start_cell)
        row, col = start_coord

        # Extract match data
        week_number = self._find_week_number(row)
        team1 = self.matchups_sheet.cell(row=row, column=col).value
        team2 = self.matchups_sheet.cell(row=row, column=col + 2).value

        matches = []

        for i in range(5):
            current_row = row + 1 + i

            # Extract match data
            team1_player = self.matchups_sheet.cell(row=current_row, column=col).value
            team2_player = self.matchups_sheet.cell(
                row=current_row, column=col + 2
            ).value

            score_cell = self.matchups_sheet.cell(row=current_row, column=col + 1)
            match_score = score_cell.value
            replay_url = score_cell.hyperlink.target if score_cell.hyperlink else None

            matches.append(
                {
                    "season": self.config.SEASON,
                    "week": week_number,
                    "team1": team1,
                    "team2": team2,
                    "team1_player": team1_player,
                    "team2_player": team2_player,
                    "match_score": match_score,
                    "replay_url": replay_url,
                }
            )

        return pd.DataFrame(matches)

    def _find_week_number(self, row: int) -> int | None:
        """Finds the week number by searching upwards from the given row."""
        week_cell_value = None
        current_row = row

        # Goes through the rows above the current row to find a cell that matches "Week X"
        while current_row > 0:
            value = self.matchups_sheet.cell(row=current_row, column=1).value

            if value and isinstance(value, str) and value.startswith("Week"):
                week_cell_value = value
                break

            current_row -= 1

        # Parse week number if found, otherwise return None
        match = re.search(r"Week\s*(\d+)", week_cell_value)
        return int(match.group(1)) if match else None
